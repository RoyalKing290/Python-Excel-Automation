# excel_automation.py
import csv
import os
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

INPUT_CSV = "sales.csv"
OUTPUT_XLSX = "report.xlsx"

def read_csv(path):
    rows = []
    with open(path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for r in reader:
            # Normalize/clean fields
            try:
                qty = float(r.get("Quantity", 0) or 0)
            except:
                qty = 0
            try:
                price = float(r.get("UnitPrice", 0) or 0)
            except:
                price = 0
            date = r.get("Date", "").strip()
            product = r.get("Product", "").strip()
            rows.append({
                "Date": date,
                "Product": product,
                "Quantity": qty,
                "UnitPrice": price,
                "Revenue": round(qty * price, 2)
            })
    return rows

def generate_report(rows, out_path):
    wb = Workbook()
    # Sheet 1: cleaned data
    ws1 = wb.active
    ws1.title = "Cleaned Sales"
    headers = ["Date", "Product", "Quantity", "UnitPrice", "Revenue"]
    ws1.append(headers)
    for r in rows:
        ws1.append([r["Date"], r["Product"], r["Quantity"], r["UnitPrice"], r["Revenue"]])

    # Auto column width simple
    for i, col in enumerate(headers, 1):
        max_len = len(col)
        for cell in ws1[get_column_letter(i)]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws1.column_dimensions[get_column_letter(i)].width = max_len + 2

    # Sheet 2: summary totals & by-product
    ws2 = wb.create_sheet(title="Summary")
    total_qty = sum(r["Quantity"] for r in rows)
    total_rev = round(sum(r["Revenue"] for r in rows), 2)

    ws2.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append([])
    ws2.append(["Total Quantity", total_qty])
    ws2.append(["Total Revenue", total_rev])
    ws2.append([])

    # Aggregate by product
    agg = defaultdict(lambda: {"qty": 0.0, "rev": 0.0})
    for r in rows:
        p = r["Product"] or "UNKNOWN"
        agg[p]["qty"] += r["Quantity"]
        agg[p]["rev"] += r["Revenue"]

    ws2.append(["Product", "Total Qty", "Total Revenue"])
    for product, v in agg.items():
        ws2.append([product, v["qty"], round(v["rev"], 2)])

    # Auto column widths for summary
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max_len + 2

    wb.save(out_path)
    print(f"✅ Report generated: {out_path}")

def main():
    if not os.path.exists(INPUT_CSV):
        print(f"Input file '{INPUT_CSV}' not found. Create it and re-run.")
        return
    rows = read_csv(INPUT_CSV)
    if not rows:
        print("No rows found in CSV. Check format.")
        return
    generate_report(rows, OUTPUT_XLSX)

if __name__ == "__main__":
    main()
